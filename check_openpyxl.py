from openpyxl import load_workbook

# Path to your Excel file
file_path = 'machine_learn_data.xlsx'

try:
    # Load the workbook
    workbook = load_workbook(file_path)
    
    # List all sheet names
    sheet_names = workbook.sheetnames
    print("Sheet names:", sheet_names)

except Exception as e:
    print(f"Error reading the Excel file: {e}")
